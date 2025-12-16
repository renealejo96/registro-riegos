from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from datetime import datetime, timedelta
import os
from supabase import create_client, Client
import csv
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import pytz

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__)

# Configurar zona horaria de Ecuador
TIMEZONE_ECUADOR = pytz.timezone('America/Guayaquil')

def get_fecha_ecuador():
    """Obtiene la fecha actual en Ecuador"""
    return datetime.now(TIMEZONE_ECUADOR).strftime('%Y-%m-%d')

def get_timestamp_ecuador():
    """Obtiene el timestamp actual en Ecuador"""
    return datetime.now(TIMEZONE_ECUADOR).isoformat()

def convertir_a_hora_ecuador(timestamp_str):
    """Convierte un timestamp UTC a hora de Ecuador"""
    try:
        # Si el timestamp ya tiene zona horaria
        if '+' in timestamp_str or timestamp_str.endswith('Z'):
            dt_utc = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            if dt_utc.tzinfo is None:
                dt_utc = pytz.utc.localize(dt_utc)
            dt_ecuador = dt_utc.astimezone(TIMEZONE_ECUADOR)
        else:
            # Si no tiene zona horaria, asumimos UTC
            dt_naive = datetime.fromisoformat(timestamp_str)
            dt_utc = pytz.utc.localize(dt_naive)
            dt_ecuador = dt_utc.astimezone(TIMEZONE_ECUADOR)
        return dt_ecuador.strftime('%H:%M:%S')
    except:
        return timestamp_str
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuración de Supabase
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')

# Inicializar cliente de Supabase
supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        print(f"✅ Conectado a Supabase: {SUPABASE_URL}")
    except Exception as e:
        print(f"⚠️  Error conectando a Supabase: {e}")
        print("La app funcionará pero sin guardar datos")
else:
    print("⚠️  Supabase no configurado - modo sin base de datos")

# Cargar módulos desde CSV
def cargar_modulos():
    modulos = []
    try:
        with open('modulos.csv', 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                modulos.append(row['modulo'])
    except Exception as e:
        print(f"Error cargando módulos: {e}")
        # Módulos por defecto si no se puede cargar el CSV
        modulos = ['11', '12', '13', '14', '15', '16', '21', '22', '23', '24']
    return sorted(modulos)

MODULOS = cargar_modulos()


@app.route('/')
def index():
    """Página principal"""
    # Usar fecha de Ecuador
    fecha_ecuador = datetime.now(TIMEZONE_ECUADOR)
    dias_espanol = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }
    meses_espanol = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
        'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
        'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    dia = dias_espanol.get(fecha_ecuador.strftime('%A'), fecha_ecuador.strftime('%A'))
    mes = meses_espanol.get(fecha_ecuador.strftime('%B'), fecha_ecuador.strftime('%B'))
    fecha_formateada = f"{dia}, {fecha_ecuador.day} de {mes} de {fecha_ecuador.year}"
    return render_template('index.html', fecha=fecha_formateada, modulos=MODULOS)


@app.route('/registrar', methods=['POST'])
def registrar_riego():
    """Registra un nuevo riego en Supabase"""
    try:
        data = request.get_json()
        print(f"📥 Datos recibidos: {data}")
        
        modulos = data.get('modulos', [])
        tipos_riego = data.get('tipos_riego', [])
        
        print(f"📋 Módulos: {modulos}")
        print(f"💧 Tipos de riego: {tipos_riego}")
        
        if not modulos:
            return jsonify({'error': 'Debe seleccionar al menos un módulo'}), 400
        
        if not tipos_riego:
            return jsonify({'error': 'Debe seleccionar al menos un tipo de riego'}), 400
        
        # Permitir pasar una fecha específica, sino usar la fecha actual de Ecuador
        fecha = data.get('fecha', get_fecha_ecuador())
        timestamp = get_timestamp_ecuador()
        
        # Preparar registros para inserción
        registros = []
        for modulo in modulos:
            for tipo_riego in tipos_riego:
                registros.append({
                    'fecha': fecha,
                    'modulo': modulo,
                    'tipo_riego': tipo_riego,
                    'timestamp': timestamp
                })
        
        print(f"✍️  Preparando {len(registros)} registros para insertar")
        
        # Insertar en Supabase
        if supabase:
            response = supabase.table('riegos').insert(registros).execute()
            print(f"✅ Respuesta de Supabase: {response}")
            
            tipos_texto = ' y '.join(['Agua' if t == 'agua' else 'Comida' for t in tipos_riego])
            
            return jsonify({
                'success': True,
                'message': f'Riego registrado: {len(modulos)} módulo(s) con {tipos_texto}',
                'registros': len(registros)
            })
        else:
            print("⚠️  Supabase no está configurado")
            return jsonify({'error': 'Supabase no está configurado'}), 500
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/registros-hoy')
def registros_hoy():
    """Obtiene los registros del día actual o de una fecha específica"""
    try:
        # Permitir pasar una fecha específica como parámetro, sino usar fecha de Ecuador
        fecha = request.args.get('fecha', get_fecha_ecuador())
        
        if supabase:
            response = supabase.table('riegos')\
                .select('*')\
                .eq('fecha', fecha)\
                .order('timestamp', desc=True)\
                .execute()
            
            registros = response.data
            
            # Formatear datos para el frontend con hora de Ecuador
            registros_formateados = []
            for reg in registros:
                hora = convertir_a_hora_ecuador(reg['timestamp'])
                
                registros_formateados.append({
                    'id': reg['id'],
                    'hora': hora,
                    'modulo': reg['modulo'],
                    'tipo_riego': 'Agua' if reg['tipo_riego'] == 'agua' else 'Comida (Fertilizante)',
                    'fecha': reg['fecha']  # Agregar fecha al resultado
                })
            
            return jsonify(registros_formateados)
        else:
            return jsonify([])
            
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/historial')
def historial():
    """Página de historial completo"""
    return render_template('historial.html')


@app.route('/resumen')
def resumen():
    """Página de resumen semanal"""
    return render_template('resumen.html', modulos=MODULOS)


@app.route('/historial-completo')
def historial_completo():
    """Obtiene todos los registros históricos"""
    try:
        if supabase:
            response = supabase.table('riegos')\
                .select('*')\
                .order('timestamp', desc=True)\
                .limit(500)\
                .execute()
            
            registros = response.data
            
            # Formatear datos
            registros_formateados = []
            for reg in registros:
                try:
                    fecha_obj = datetime.strptime(reg['fecha'], '%Y-%m-%d')
                    fecha = fecha_obj.strftime('%d/%m/%Y')
                    hora = datetime.fromisoformat(reg['timestamp']).strftime('%H:%M:%S')
                except:
                    fecha = reg['fecha']
                    hora = reg['timestamp']
                
                registros_formateados.append({
                    'fecha': fecha,
                    'hora': hora,
                    'modulo': reg['modulo'],
                    'tipo_riego': 'Agua' if reg['tipo_riego'] == 'agua' else 'Comida (Fertilizante)'
                })
            
            return jsonify(registros_formateados)
        else:
            return jsonify([])
            
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/eliminar/<int:id>', methods=['DELETE'])
def eliminar_riego(id):
    """Elimina un registro de riego"""
    try:
        if supabase:
            response = supabase.table('riegos').delete().eq('id', id).execute()
            print(f"🗑️  Registro {id} eliminado")
            
            return jsonify({
                'success': True,
                'message': 'Registro eliminado correctamente'
            })
        else:
            return jsonify({'error': 'Supabase no está configurado'}), 500
            
    except Exception as e:
        print(f"❌ Error al eliminar: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/editar/<int:id>', methods=['PUT'])
def editar_riego(id):
    """Edita un registro de riego"""
    try:
        data = request.get_json()
        modulo = data.get('modulo')
        tipo_riego = data.get('tipo_riego')
        
        if not modulo or not tipo_riego:
            return jsonify({'error': 'Datos incompletos'}), 400
        
        if supabase:
            response = supabase.table('riegos')\
                .update({'modulo': modulo, 'tipo_riego': tipo_riego})\
                .eq('id', id)\
                .execute()
            
            print(f"✏️  Registro {id} actualizado: {modulo} - {tipo_riego}")
            
            return jsonify({
                'success': True,
                'message': 'Registro actualizado correctamente'
            })
        else:
            return jsonify({'error': 'Supabase no está configurado'}), 500
            
    except Exception as e:
        print(f"❌ Error al editar: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/estadisticas')
def estadisticas():
    """Obtiene estadísticas de riegos"""
    try:
        if supabase:
            # Registros de hoy (fecha de Ecuador)
            fecha_hoy = get_fecha_ecuador()
            hoy = supabase.table('riegos')\
                .select('*', count='exact')\
                .eq('fecha', fecha_hoy)\
                .execute()
            
            # Total de registros
            total = supabase.table('riegos')\
                .select('*', count='exact')\
                .execute()
            
            return jsonify({
                'registros_hoy': hoy.count if hoy.count else 0,
                'total_registros': total.count if total.count else 0
            })
        else:
            return jsonify({'registros_hoy': 0, 'total_registros': 0})
            
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/resumen-semanal')
def resumen_semanal():
    """Obtiene resumen semanal de riegos"""
    try:
        # Diccionario para traducir días
        dias_espanol = {
            'Monday': 'Lunes',
            'Tuesday': 'Martes',
            'Wednesday': 'Miércoles',
            'Thursday': 'Jueves',
            'Friday': 'Viernes',
            'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        
        semana = request.args.get('semana')  # formato: 2025-50
        
        if not semana:
            # Si no se especifica, usar la semana actual de Ecuador
            hoy = datetime.now(TIMEZONE_ECUADOR)
            year, week_num, _ = hoy.isocalendar()
            semana = f"{year}-{week_num:02d}"
        
        # Parsear año y semana
        year, week_num = map(int, semana.split('-'))
        
        # Calcular fecha de inicio y fin de la semana
        first_day_of_year = datetime(year, 1, 1)
        days_to_monday = (first_day_of_year.weekday()) % 7
        first_monday = first_day_of_year - timedelta(days=days_to_monday)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)
        
        fecha_inicio = week_start.strftime('%Y-%m-%d')
        fecha_fin = week_end.strftime('%Y-%m-%d')
        
        if supabase:
            response = supabase.table('riegos')\
                .select('*')\
                .gte('fecha', fecha_inicio)\
                .lte('fecha', fecha_fin)\
                .execute()
            
            registros = response.data
            
            # Procesar datos para el resumen
            resumen = {}
            for reg in registros:
                fecha_obj = datetime.strptime(reg['fecha'], '%Y-%m-%d')
                dia_ingles = fecha_obj.strftime('%A')
                dia_espanol = dias_espanol.get(dia_ingles, dia_ingles)
                dia = f"{dia_espanol} {fecha_obj.strftime('%d/%m')}"
                modulo = reg['modulo']
                tipo = reg['tipo_riego']
                
                key = f"{fecha_obj.strftime('%Y-%m-%d')}|{modulo}"
                
                if key not in resumen:
                    resumen[key] = {
                        'fecha': reg['fecha'],
                        'dia': dia,
                        'modulo': modulo,
                        'agua': False,
                        'comida': False,
                        'semana': f"{year}{week_num:02d}"
                    }
                
                if tipo == 'agua':
                    resumen[key]['agua'] = True
                elif tipo == 'comida':
                    resumen[key]['comida'] = True
            
            # Convertir a lista y ordenar
            resultado = sorted(resumen.values(), key=lambda x: (x['fecha'], x['modulo']))
            
            return jsonify({
                'semana': semana,
                'datos': resultado
            })
        else:
            return jsonify({'semana': semana, 'datos': []})
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/exportar-excel')
def exportar_excel():
    """Exporta el resumen semanal a Excel"""
    try:
        # Diccionario para traducir días
        dias_espanol = {
            'Monday': 'Lunes',
            'Tuesday': 'Martes',
            'Wednesday': 'Miércoles',
            'Thursday': 'Jueves',
            'Friday': 'Viernes',
            'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        
        semana = request.args.get('semana')
        
        if not semana:
            hoy = datetime.now(TIMEZONE_ECUADOR)
            year, week_num, _ = hoy.isocalendar()
            semana = f"{year}-{week_num:02d}"
        
        # Obtener datos del resumen
        year, week_num = map(int, semana.split('-'))
        first_day_of_year = datetime(year, 1, 1)
        days_to_monday = (first_day_of_year.weekday()) % 7
        first_monday = first_day_of_year - timedelta(days=days_to_monday)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)
        
        fecha_inicio = week_start.strftime('%Y-%m-%d')
        fecha_fin = week_end.strftime('%Y-%m-%d')
        
        if supabase:
            response = supabase.table('riegos')\
                .select('*')\
                .gte('fecha', fecha_inicio)\
                .lte('fecha', fecha_fin)\
                .execute()
            
            registros = response.data
            
            # Crear Excel
            wb = Workbook()
            ws = wb.active
            ws.title = f"Semana {week_num}"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Encabezados
            headers = ['Semana Año', 'Día', 'Módulo', 'Agua', 'Comida']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Procesar datos
            resumen = {}
            for reg in registros:
                fecha_obj = datetime.strptime(reg['fecha'], '%Y-%m-%d')
                dia_ingles = fecha_obj.strftime('%A')
                dia_espanol_nom = dias_espanol.get(dia_ingles, dia_ingles)
                dia = f"{dia_espanol_nom} {fecha_obj.strftime('%d/%m')}"
                modulo = reg['modulo']
                tipo = reg['tipo_riego']
                
                key = f"{fecha_obj.strftime('%Y-%m-%d')}|{modulo}"
                
                if key not in resumen:
                    resumen[key] = {
                        'fecha': reg['fecha'],
                        'dia': dia,
                        'modulo': modulo,
                        'agua': False,
                        'comida': False,
                        'semana': f"{year}{week_num:02d}"
                    }
                
                if tipo == 'agua':
                    resumen[key]['agua'] = True
                elif tipo == 'comida':
                    resumen[key]['comida'] = True
            
            # Escribir datos
            datos = sorted(resumen.values(), key=lambda x: (x['fecha'], x['modulo']))
            
            for idx, item in enumerate(datos, 2):
                ws.cell(row=idx, column=1, value=item['semana'])
                ws.cell(row=idx, column=2, value=item['dia'])
                ws.cell(row=idx, column=3, value=item['modulo'])
                ws.cell(row=idx, column=4, value='✓' if item['agua'] else '✗')
                ws.cell(row=idx, column=5, value='✓' if item['comida'] else '✗')
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            
            # Guardar en memoria
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'resumen_riegos_semana_{week_num}_{year}.xlsx'
            )
        else:
            return jsonify({'error': 'Supabase no está configurado'}), 500
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
